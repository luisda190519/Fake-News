from ctypes import sizeof
import re
from tkinter.ttk import Sizegrip
import urllib.request
from bs4 import BeautifulSoup as bs
from datetime import datetime
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from itertools import chain
import numpy as np
import openpyxl
import pandas as pd
from wordcloud import WordCloud, STOPWORDS
import matplotlib.pyplot as plt

# descomentar esto si te da algun error relacionado a ellas
# nltk.download('stopwords')
# nltk.download('punkt')

def remove_values_from_list(the_list, val):
   return [value for value in the_list if value != val]

# se utiliza para limpiar los valores que no sea alfanumericos
def limpia_alfanum(texto):

    return re.compile(r'\W+', re.UNICODE).split(texto)

# funcion que devuelve el numero del mes    
def numero_mes(mes):
    meses=["enero", "febrero", "marzo", "abril", "mayo", "junio","julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    if(len(str(meses.index(mes) + 1)) == 1):
        return '0' + str(meses.index(mes) + 1)
    else:
        return str(meses.index(mes) + 1)

# funcion que devuelve la fecha en formato dd/mm/aaaa
def limpiar_fechas(fecha):
    now = datetime.now()
    if(len(fecha)<=5):
        return (str(now.day)+'/'+str(now.date())[5:7]+'/'+str(now.year))
    else:
        fecha_limpia=fecha[5:].split()
    return (str(fecha_limpia[0])+'/'+numero_mes(str(fecha_limpia[1]))+'/'+str(fecha_limpia[2]))

# funcion que genera la distribucion de freciencias
def distribucion_frecuencias(palabras):
    frec=[]
    for i in range(len(palabras)):
        if palabras[i] not in palabras[:i]:
            
            frec.append([palabras[i], palabras.count(palabras[i])])
    frec.sort(key=lambda tit: tit[1], reverse=True)
    return frec

# funcion que genera la nube de palabras
def word_cloud(palabras):
    palabras=str(" ".join(palabras))
    wordcloud = WordCloud(width = 600, height = 600,
                background_color ='white',
                min_font_size = 10).generate(palabras)
 
    # plot the WordCloud image                      
    plt.figure(figsize = (6, 6), facecolor = None)
    plt.imshow(wordcloud)
    plt.axis("off")
    plt.tight_layout(pad = 0)
    plt.savefig("WordCloud")
    plt.show()

#función que genera n-gramas
def n_gramas(words,n):
    ngramas = []
    for i in range(len(words)-(n-1)):
        if not words[i:i+n] in ngramas:#verifica que no se repitan los n-gramas
            ngramas.append(words[i:i+n])
    return ngramas

def grafico_barras(eje_x, eje_y, nombre, lista_x, lista_y, rotacion):
    fig, ax = plt.subplots()
    #Colocamos una etiqueta en el eje Y
    ax.set_ylabel(str(eje_y))
    #Colocamos una etiqueta en el eje X
    ax.set_xlabel(str(eje_x))   
    fig.set_figheight(6)
    fig.set_figwidth(11)
    #Creamos la grafica de barras utilizando 'paises' como eje X y 'ventas' como eje y.
    plt.subplots_adjust(bottom=0.4)
    plt.bar(lista_x, lista_y,width=0.5)
    plt.xticks(rotation=rotacion)
    plt.title(str(nombre))
    plt.savefig(str(nombre))
    #Finalmente mostramos la grafica con el metodo show()
    plt.show()

if __name__ == "__main__":
    contp=1
    n=int(input("digite n "))
    while(n>900):
        n=int(input("digite n (menor a 900)"))
    fecha=input("digite la fecha en formato dd/mm/aaaa ")
    numt=0
    lista_titulos=[]
    while(numt<n):# Mientras que no se han obtenido los n titulos
        url="https://www.bbc.com/mundo/topics/c7zp57yyz25t/page/"+str(contp)
        datos=urllib.request.urlopen(url).read().decode()
        soup=bs(datos, 'html.parser')
        
        #se buscan todos los titulos
        titulos = soup.find_all('span', class_="lx-stream-post__header-text gs-u-align-middle")
        #Se buscan todas las fechas
        fechas = soup.find_all('span', class_='qa-post-auto-meta')
        for i in range(len(titulos)):
            lista_titulos.append([titulos[i].text, limpiar_fechas(str(fechas[i].text))])
            numt+=1
            if(numt>=n):
                break
        contp+=1
    
    #se guardan los titulos con sus fechas en un archivo de excel 
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Titulos"
    # Crea la fila del encabezado con los títulos
    hoja.append(('Titulo', 'fecha'))
    for tit in lista_titulos:
        hoja.append(tit)
    
    # Se hace la limpieza de valores que no sean alfanumericos
    titulos_alfanumericos=[]
    for i in range(n):
        limp=limpia_alfanum(lista_titulos[i][0])
        if('' in limp):
            limp=remove_values_from_list(limp,'')
        titulos_alfanumericos.append(limp)
    for i in range(n):
        lista_titulos[i][0]=titulos_alfanumericos[i]
    
    #se añade otra hoja con los titulos solo con valores alfanumericos
    hoja2 = wb.create_sheet("Titulos alfanum")
    wb.active=hoja2
    hoja2.append(('Titulos solo con valores alfanumericos', 'fecha'))
    for tit in lista_titulos:
        tl=''
        for i in range(len(tit[0])):
            tl+=str(tit[0][i])+ (' ' if i< len(tit[0])-1 else '')
        hoja2.append([tl,tit[1]])
    
    titulares_fake=[]
    for k in range(30):
        pair={}
        words=[]
        for i in range(len(lista_titulos)):
            for pal in lista_titulos[i][0]:
                words.append(pal)#se añade cada palabra de los titulares a una lista
        
        def make(words):
            for i in range(len(words)-1):
                yield(words[i],words[i+1])#se forman parejas de palabras para que se cumpla que sean cadenas de markov

        pair=make(words)
        #Se crea un diccionario con los pares de palabras y se verifica que la primera palabra de dicho par no sea ya una clave del diccionario
        word_dict = {}
        for word_1, word_2 in pair:
            if word_1 in word_dict.keys():
                word_dict[word_1].append(word_2)
            else:
                word_dict[word_1] = [word_2]
        word_dict[words[-1]]=[words[-2]]
        while(True):#se valida que la primera palabra del titular generado sea mayúscula
            first_word = np.random.choice(words)
            if(first_word[0].isupper() and len(first_word)>0):
                break

        while first_word.islower()==False:#se crea una cadena de markov 
            chain = [first_word]
            n_words = 10
            first_word = np.random.choice(words)
            
            for i in range(n_words):
                palabra=np.random.choice(word_dict[chain[-1]])
                chain.append(palabra)
            titulares_fake.append(' '.join(chain))
    print("\n\nTitulares fake")
    print(titulares_fake)
    #Aqui se hace la limpieza de las stopWords. Tomado de https://www.geeksforgeeks.org/removing-stop-words-nltk-python/
    stop_words = set(stopwords.words('spanish'))
    titulos_sin_stopwords=[]
    for i in range (n):
        word_tokens=[]
        for j in range(len(lista_titulos[i][0])):
            word_tokens.append(lista_titulos[i][0][j].lower())    
        filtered_sentence = [w for w in word_tokens if not w.lower() in stop_words]
        filtered_sentence = []
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence.append(w)
        titulos_sin_stopwords.append(filtered_sentence)

    #se añade otra hoja de excel con los valores limpios(sin stop words)
    hoja3 = wb.create_sheet("Titulos sin stop words")
    wb.active=hoja3
    hoja3.append(('Titulos limpios', 'fecha'))
    for i in range(len(titulos_sin_stopwords)):
        hoja3.append([" ".join(titulos_sin_stopwords[i]),lista_titulos[i][1]])
    
    #se añade otra hoja de excel con los 30 titulares fake generados
    hoja4 = wb.create_sheet("Titulos sin stop words")
    wb.active=hoja4
    hoja4.append(['Titulares fake'])
    for tit in titulares_fake:
        hoja4.append([tit])
    wb.save('TitulosFakeNews.xlsx')

    #se realiza la distribucion de frecuencias
    palabras=[]
    for i in range(len(titulos_sin_stopwords)):
        for pal in titulos_sin_stopwords[i]:
            palabras.append(pal)
    print("\n\nDistribucion de frecuencias\n")
    distr=distribucion_frecuencias(palabras)[:19]
    print(distr)
    ej_x=[]
    ej_y=[]
    for i in range (len(distr) if len(distr)<10 else 10 ):
        ej_x.append(distr[i][0])
        ej_y.append(distr[i][1])
    grafico_barras("Palabras", "cantidad", "Distrubucion de frecuencia", ej_x, ej_y, 0)

    #se realiza la distribucion de frecuencias en la fecha dada
    palabrasfecha=[]
    for i in range(len(lista_titulos)):
        if(lista_titulos[i][1]==fecha):
            for pal in titulos_sin_stopwords[i]:
                palabrasfecha.append(pal)
    print("\n\nPalabras mas usadas en la fecha ",fecha)
    print(distribucion_frecuencias(palabrasfecha)[:9])

    #Bigramas y trigramas
    bigramas=n_gramas(palabras,2)
    trigramas=n_gramas(palabras,3)
    bigr=[]
    trig=[]
    for bigrama in bigramas:
        frecuencia=0
        for i in range(len(palabras)-1):
            if bigrama[0] in palabras[i] and bigrama[1] in palabras[i+1]:
                frecuencia+=1
        bigr.append([str(bigrama[0])+" "+str(bigrama[1]),frecuencia])#se crean los bigramas y se guardan en una lista junto con su frecuencia
    for trigrama in trigramas:
        frecuencia=0
        for i in range(len(palabras)-2):
            if trigrama[0] in palabras[i] and trigrama[1] in palabras[i+1] and trigrama[2] in palabras[i+2]:
                frecuencia+=1
        trig.append([str(trigrama[0])+" "+str(trigrama[1])+" "+trigrama[2],frecuencia])
    bigramas_orden=sorted(bigr,key=lambda bigrama:bigrama[1],reverse=True)
    trigramas_orden=sorted(trig,key=lambda trigrama:trigrama[1],reverse=True)
    print("\n10 bigramas más observados: ")
    for i in range(10):
        print(bigramas_orden[i])
    print("\n10 trigramas más observados: ")
    for i in range(10):
        print(trigramas_orden[i])
    
    word_cloud(palabras)

    fech=[]
    for i in range(len(lista_titulos)):
        fech.append([str(lista_titulos[i][1])[3:5],str(lista_titulos[i][1])[6:]])
    ej_y=[]
    print("\n\nNoticias por mes")
    for i in range(12):        
        ej_y.append(fech.count([('0' if i+1<10 else '')+str(i+1),'2022']))
        print("Mes", i+1,": ", ej_y[i])
    grafico_barras("Mes", "Cantidad", "Noticias por mes", list(range(1,12+1)), ej_y, 0)

    ej_x=[]
    ej_y=[]
    for i in range(10):
        ej_x.append(bigramas_orden[i][0])
        ej_y.append(bigramas_orden[i][1])
    grafico_barras("Bigrama", "Cantidad", "Distribucion de bigramas", ej_x, ej_y, 90)

    ej_x=[]
    ej_y=[]
    for i in range(10):
        ej_x.append(trigramas_orden[i][0])
        ej_y.append(trigramas_orden[i][1])
    grafico_barras("Trigrama", "Cantidad", "Distribucion de trigramas", ej_x, ej_y, 90)